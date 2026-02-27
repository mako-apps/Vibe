"""
Vibe Document Renderer — lightweight HTTP service for PDF/PNG/XLSX generation.

Endpoints:
  POST /render       — render HTML to PDF or PNG
  POST /xlsx         — generate styled XLSX from columns + rows JSON
  GET  /health       — health check
"""

import io
import json
import logging
import sys
from flask import Flask, request, jsonify, send_file
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
logging.basicConfig(stream=sys.stderr, level=logging.INFO,
                    format="[DocRenderer] %(message)s")
log = logging.getLogger(__name__)

# ── Shared styles ──

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)

# ── HTML template for PDF/PNG export ──

EXPORT_HTML_TEMPLATE = """<!DOCTYPE html>
<html dir="rtl" lang="fa">
<head>
<meta charset="UTF-8"/>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
@page {{ size: A4; margin: 15mm; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Tahoma, Arial, sans-serif;
  font-size: 13px;
  line-height: 1.5;
  color: #1a1a1a;
  direction: rtl;
  text-align: right;
}}
h1 {{
  font-size: 20px;
  font-weight: 700;
  color: #2B5797;
  margin-bottom: 12px;
  padding-bottom: 8px;
  border-bottom: 2px solid #2B5797;
}}
.meta {{
  font-size: 11px;
  color: #666;
  margin-bottom: 14px;
}}
table {{
  width: 100%;
  border-collapse: collapse;
  margin-top: 10px;
}}
th {{
  background: #2B5797;
  color: #fff;
  font-weight: 600;
  padding: 8px 10px;
  text-align: right;
  border: 1px solid #1e3f6f;
  white-space: nowrap;
}}
td {{
  padding: 6px 10px;
  border: 1px solid #d0d0d0;
  text-align: right;
  vertical-align: top;
}}
tr:nth-child(even) td {{ background: #f5f7fa; }}
</style>
</head>
<body>
<h1>{title}</h1>
<div class="meta">{meta}</div>
<table>
<thead><tr>{header_cells}</tr></thead>
<tbody>
{body_rows}
</tbody>
</table>
</body>
</html>"""


def _escape(text):
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def _build_export_html(title, columns, rows, meta=""):
    header_cells = "".join(f"<th>{_escape(c)}</th>" for c in columns)
    body_rows = ""
    for row in rows:
        cells = "".join(f"<td>{_escape(v)}</td>" for v in row)
        body_rows += f"<tr>{cells}</tr>\n"
    return EXPORT_HTML_TEMPLATE.format(
        title=_escape(title),
        meta=_escape(meta),
        header_cells=header_cells,
        body_rows=body_rows,
    )


# ── Routes ──

@app.route("/health", methods=["GET"])
def health():
    return jsonify(status="ok"), 200


@app.route("/render", methods=["POST"])
def render():
    """Render HTML string or structured data to PDF or PNG."""
    data = request.get_json(force=True)
    fmt = data.get("format", "pdf").lower()
    title = data.get("title", "Document")

    # Accept either raw html or structured columns+rows
    html_content = data.get("html")
    if not html_content:
        columns = data.get("columns", [])
        rows = data.get("rows", [])
        meta = data.get("meta", "")
        if not columns:
            return jsonify(error="Provide 'html' or 'columns'+'rows'"), 400
        html_content = _build_export_html(title, columns, rows, meta)

    try:
        html_obj = HTML(string=html_content)
        buf = io.BytesIO()

        if fmt == "png":
            html_obj.write_png(buf)
            mimetype = "image/png"
            ext = "png"
        else:
            html_obj.write_pdf(buf)
            mimetype = "application/pdf"
            ext = "pdf"

        buf.seek(0)
        log.info(f"Rendered {fmt} ({buf.getbuffer().nbytes} bytes)")
        return send_file(buf, mimetype=mimetype,
                         download_name=f"{title}.{ext}", as_attachment=False)
    except Exception as e:
        log.error(f"Render failed: {e}")
        return jsonify(error=str(e)), 500


@app.route("/xlsx", methods=["POST"])
def xlsx():
    """Generate a styled XLSX file from columns + rows."""
    data = request.get_json(force=True)
    columns = data.get("columns", [])
    rows = data.get("rows", [])
    title = data.get("title", "Spreadsheet")
    sheet_rtl = data.get("rtl", True)

    if not columns:
        return jsonify(error="'columns' is required"), 400

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.sheet_view.rightToLeft = sheet_rtl

        # Column widths
        for i, col in enumerate(columns, 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = max(len(str(col)) + 6, 18)

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = HEADER_ALIGN

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                if col_idx > len(columns):
                    break
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = DATA_ALIGN

            # Fill empty cells with border
            for col_idx in range(len(row_data) + 1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.border = THIN_BORDER
                cell.alignment = DATA_ALIGN

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        log.info(f"Generated XLSX: {len(columns)} cols, {len(rows)} rows, {buf.getbuffer().nbytes} bytes")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            download_name=f"{title}.xlsx",
            as_attachment=False,
        )
    except Exception as e:
        log.error(f"XLSX generation failed: {e}")
        return jsonify(error=str(e)), 500


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5050
    log.info(f"Starting doc renderer on port {port}")
    app.run(host="127.0.0.1", port=port, debug=False)
